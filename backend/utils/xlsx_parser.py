"""
XLSX Parser — ZR en Artex formaten
Detecteert leverancier op basis van kolomkoppen en parseert naar BMS-formaat.
"""

import openpyxl
import re


def _clean(val) -> str:
    if val is None:
        return ''
    return str(val).strip()


def _missing_fields(item: dict) -> list[str]:
    missing = []
    if not item.get('width_cm'):
        missing.append('breedte')
    if not item.get('composition'):
        missing.append('samenstelling')
    if not item.get('price_rrp') and not item.get('price_excl'):
        missing.append('prijs')
    return missing


# ─── ZR XLSX ──────────────────────────────────────────────────────────────────

ZR_COL_MAP = {
    'item no.':           'item_no',
    'item no':            'item_no',
    'item':               'article',
    'variant':            'variant',
    'purchase price':     'price_excl',
    'excl.':              'price_excl',
    'rrp':                'price_rrp',
    'incl. vat':          'price_rrp',
    'width':              'width_cm',
    'width cm':           'width_cm',
    'gram':               'weight_g',
    'no. colors':         'colors',
    'no colors':          'colors',
    'composition %':      'composition',
    'composition':        'composition',
    'repeat \nheight':    'repeat_h_cm',
    'repeat \nwidth':     'repeat_w_cm',
    'repeat height':      'repeat_h_cm',
    'repeat width':       'repeat_w_cm',
    'additional':         'features',
}


def _map_zr_header(headers: list) -> dict:
    """Geeft {col_index: field_name} mapping voor ZR formaat."""
    mapping = {}
    for i, h in enumerate(headers):
        if h is None:
            continue
        key = str(h).lower().strip().replace('\n', ' ')
        for pattern, field in ZR_COL_MAP.items():
            if key.startswith(pattern.lower()):
                if field not in mapping.values():  # eerste match wint
                    mapping[i] = field
                break
    return mapping


def parse_zr_xlsx(path: str) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return _empty_result()

    # Eerste niet-lege rij = header
    header_row = None
    header_idx = 0
    for i, row in enumerate(rows):
        if any(c is not None for c in row):
            header_row = row
            header_idx = i
            break

    if header_row is None:
        return _empty_result()

    col_map = _map_zr_header(list(header_row))

    items = []
    for row in rows[header_idx + 1:]:
        if not any(c is not None for c in row):
            continue

        item = {}
        for col_idx, field in col_map.items():
            if col_idx < len(row):
                item[field] = _clean(row[col_idx])

        item_no = item.get('item_no', '')
        if not item_no or not re.match(r'^\d+', item_no):
            continue

        # Normaliseer
        item['item_no'] = item_no
        item['repeat_h_cm'] = item.get('repeat_h_cm', '')
        item['repeat_w_cm'] = item.get('repeat_w_cm', '')
        item['martindale'] = None
        item['flame_retardant'] = 'FR' in item.get('features', '').upper() or \
                                   'FR' in item.get('composition', '').upper()
        item['coating_cs'] = 'CS' in item.get('composition', '').upper()
        item['missing_fields'] = _missing_fields(item)

        # Verwijder 0-waarden voor repeat
        if item.get('repeat_h_cm') in ('0', '0.0'):
            item['repeat_h_cm'] = ''
        if item.get('repeat_w_cm') in ('0', '0.0'):
            item['repeat_w_cm'] = ''

        items.append(item)

    return _build_result(items)


# ─── ARTEX XLSX ───────────────────────────────────────────────────────────────

ARTEX_COL_MAP = {
    'naam':                  'article',
    'hoogte / breedte':      'width_cm',
    'hoogte/breedte':        'width_cm',
    'rapport ca. cm':        'repeat_h_cm',
    'rapport':               'repeat_h_cm',
    'gewicht gr/m1':         'weight_g',
    'gewicht':               'weight_g',
    'wasadvies':             'wash_advice',
    'samenstelling in %':    'composition',
    'samenstelling':         'composition',
    'vlamvertragend':        'flame_retardant_raw',
    'kamerhoog':             'full_height',
    'verkoopprijs incl':     'price_rrp',
    'verkoopprijs':          'price_rrp',
}


def _map_artex_header(headers: list) -> dict:
    mapping = {}
    for i, h in enumerate(headers):
        if h is None:
            continue
        key = str(h).lower().strip()
        for pattern, field in ARTEX_COL_MAP.items():
            if key.startswith(pattern.lower()):
                if field not in mapping.values():
                    mapping[i] = field
                break
    return mapping


def parse_artex_xlsx(path: str) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)

    all_items = []
    seen_names = set()  # Artex heeft geen unieke artikel nummers

    for sheet_name in wb.sheetnames:
        # Sla alfabet-overzicht sheet over
        if 'alfabet' in sheet_name.lower() or 'overzicht' in sheet_name.lower():
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Zoek header rij
        header_row = None
        header_idx = 0
        for i, row in enumerate(rows[:5]):
            if row[0] is not None and str(row[0]).lower().strip() in ('naam', 'name', 'artikel'):
                header_row = row
                header_idx = i
                break
            if row[0] is not None and i == 0:
                header_row = row
                header_idx = i

        if header_row is None:
            continue

        col_map = _map_artex_header(list(header_row))
        if not col_map:
            continue

        for row_idx, row in enumerate(rows[header_idx + 1:], header_idx + 1):
            if not any(c is not None for c in row):
                continue

            item = {}
            for col_idx, field in col_map.items():
                if col_idx < len(row):
                    item[field] = _clean(row[col_idx])

            name = item.get('article', '')
            if not name or name.lower() in ('naam', 'name'):
                continue

            # Unieke sleutel: naam + sheet
            key = f"{sheet_name}::{name}"
            if key in seen_names:
                continue
            seen_names.add(key)

            # Artex heeft geen artikelnummers — gebruik naam als ID
            item['item_no'] = f"{sheet_name[:3].upper()}-{name}"
            item['collection'] = sheet_name
            item['repeat_h_cm'] = item.get('repeat_h_cm', '')
            item['repeat_w_cm'] = ''
            item['martindale'] = None
            item['price_excl'] = ''

            # Vlamvertragend
            fr_raw = item.get('flame_retardant_raw', '')
            cs_in_comp = 'trevira cs' in item.get('composition', '').lower() or \
                         'cs' in item.get('composition', '').lower()
            item['flame_retardant'] = bool(fr_raw and fr_raw.strip() not in ('', '-', 'None'))
            item['coating_cs'] = cs_in_comp
            item['missing_fields'] = _missing_fields(item)

            all_items.append(item)

    return _build_result(all_items)


# ─── EIJFFINGER PDF ───────────────────────────────────────────────────────────

def parse_eijffinger_pdf(path: str) -> dict:
    """
    Eijffinger behang-pricelist — twee kolommen per pagina.

    Structuur per pagina:
      - Collectienaam (AMBER, ANAM, …) staat gecentreerd op y ≈ 44
      - Header op y ≈ 96–107: Item No. | Roll size width/length | Price per roll | Price per m2
      - Datarijen: 6-cijferig item_no links (~x38) of rechts (~x293)

    Exacte x-posities uit de PDF:
      Links:  item_no≈38  width≈91  length≈131  price_roll≈170  price_m2≈231
      Rechts: item_no≈293 width≈346 length≈386  price_roll≈424  price_m2≈486

    Murals-sectie (y>400 op dezelfde pagina): alleen item_no + prijs per item, geen breedte.
    Footer (y>480): telefoon/email — overslaan.
    """
    import pdfplumber

    ITEM_RE = re.compile(r'^\d{6}$')

    # Kolomgrenzen per helft (gebaseerd op werkelijke x-posities)
    LEFT_COLS = {
        'item_no':    (0,   85),
        'width_cm':   (85,  125),
        'length_m':   (125, 167),
        'price_roll': (167, 228),   # € + getal op ~170/181
        'price_m2':   (228, 292),   # € + getal op ~231/238
    }
    RIGHT_COLS = {
        'item_no':    (290, 340),
        'width_cm':   (340, 382),
        'length_m':   (382, 422),
        'price_roll': (422, 484),   # € + getal op ~424/436
        'price_m2':   (484, 540),   # € + getal op ~486/493
    }

    # Murals: alleen item_no + prijs per item (geen breedte/lengte)
    LEFT_MURAL  = {'item_no': (0, 155),   'price_item': (155, 292)}
    RIGHT_MURAL = {'item_no': (290, 410), 'price_item': (410, 540)}

    SKIP_WORDS = {
        'Item', 'No.', 'Roll', 'size', 'Recommended', 'Retail', 'Price*',
        'width', 'cm', 'length', 'm', 'per', 'roll', 'm2', 'item',
        'PRICELIST', 'PRIJSLIJST', 'PREISLISTE', 'LISTE', 'DE', 'PRIX',
        '*Recommended', '*De', '*Die', '*Les',
    }

    def _to_float(tokens: list) -> float | None:
        s = ' '.join(tokens).replace('€', '').replace(',', '.').strip()
        # Pak het eerste getal dat eruit ziet als een prijs
        m = re.search(r'\d+\.\d{2}', s)
        if m:
            try:
                return float(m.group())
            except ValueError:
                pass
        return None

    def _classify(x: float, cols: dict) -> str | None:
        for field, (x0, x1) in cols.items():
            if x0 <= x < x1:
                return field
        return None

    def _parse_rows(words: list, cols: dict, y_max: float = 9999) -> list[dict]:
        """
        Groepeer woorden per y-rij en extraheer items.

        Gebruikt proximity-grouping (tolerance 3px) in plaats van vaste buckets,
        omdat pdfplumber consistente y-offsets van ~0.47px toont tussen item_no
        en prijswoorden op dezelfde visuele rij. Vaste buckets (/6, /8) falen
        afhankelijk van de absolute y-positie.
        """
        TOLERANCE = 3.0  # px — groter dan max offset (~0.47), kleiner dan rij-afstand (~11px)

        # Filter en sorteer op y
        filtered = [w for w in words if w['top'] <= y_max and w['text'] not in SKIP_WORDS]
        filtered.sort(key=lambda w: w['top'])

        # Groepeer: elke nieuw woord gaat naar de dichtstbijzijnde bestaande rij
        # binnen tolerance, anders een nieuwe rij.
        rows: dict[float, list] = {}
        for w in filtered:
            y = w['top']
            if rows:
                closest = min(rows.keys(), key=lambda ry: abs(ry - y))
                if abs(closest - y) <= TOLERANCE:
                    rows[closest].append(w)
                    continue
            rows[y] = [w]

        items = []
        for y in sorted(rows.keys()):
            line = sorted(rows[y], key=lambda w: w['x0'])
            if not line:
                continue
            first = line[0]
            if not ITEM_RE.match(first['text']):
                continue
            raw: dict[str, list] = {f: [] for f in cols}
            for w in line:
                f = _classify(w['x0'], cols)
                if f:
                    raw[f].append(w['text'])
            items.append(raw)

        return items

    all_items: list[dict] = []
    current_coll = ''

    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            words = page.extract_words()
            if not words:
                continue

            page_height = page.height or 842

            # ── Detecteer collectienaam ────────────────────────────────────
            # Staat gecentreerd op y ≈ 44, vóór de headerrij op y ≈ 96
            for w in words:
                if 30 <= w['top'] <= 75 and w['x0'] > 100:
                    txt = w['text'].strip()
                    # Collectienaam = niet-numeriek, niet in SKIP_WORDS, minstens 2 tekens
                    # (len>1 filter voorkomt losse letters zoals 'P' uit gespacieerde koppen)
                    if txt and len(txt) > 1 and txt not in SKIP_WORDS and not txt.isdigit():
                        current_coll = txt
                        break

            # ── Footer-grens: vanaf telefoon/email-regels (~y 770) ──────────
            footer_y = page_height - 80  # laatste 80px overslaan

            # ── WALLPOWER FAVOURITES: 3-koloms mural-overzichtspagina ────────
            # Detectie: 'WALLPOWER' tekst op y≈61. Items staan in 3 kolommen
            # (x≈52, x≈206, x≈355) met prijs recht ernaast — géén breedte/lengte.
            if any(w['text'] == 'WALLPOWER' and 55 <= w['top'] <= 70 for w in words):
                WP_SPECS = [
                    ({'item_no': (0,   115), 'price_item': (115, 200)},
                     lambda w: w['x0'] < 200),
                    ({'item_no': (200, 265), 'price_item': (265, 350)},
                     lambda w: 200 <= w['x0'] < 350),
                    ({'item_no': (350, 420), 'price_item': (420, 545)},
                     lambda w: w['x0'] >= 350),
                ]
                for col_spec, col_filter in WP_SPECS:
                    col_words = [w for w in words if col_filter(w) and w['top'] > 100]
                    for raw in _parse_rows(col_words, col_spec, footer_y):
                        item_no = ' '.join(raw.get('item_no', []))
                        if not ITEM_RE.match(item_no):
                            continue
                        price_f = _to_float(raw.get('price_item', []))
                        missing = [] if price_f else ['prijs']
                        all_items.append({
                            'item_no':         item_no,
                            'article':         'WALLPOWER',
                            'collection':      'WALLPOWER',
                            'width_cm':        '',
                            'length_m':        '',
                            'repeat_h_cm':     '',
                            'repeat_w_cm':     '',
                            'composition':     '',
                            'martindale':      None,
                            'flame_retardant': False,
                            'coating_cs':      False,
                            'price_excl':      price_f,
                            'price_rrp':       None,
                            'colors':          '',
                            'features':        'mural',
                            'missing_fields':  missing,
                        })
                continue  # Sla normale parsing over voor WALLPOWER-pagina

            # ── Murals-sectie: detecteer "MURALS" koptekst ─────────────────
            mural_y_start = None
            for w in words:
                if 'MURALS' in w['text'].upper() and w['top'] > 200:
                    mural_y_start = w['top']
                    break

            # ── Normale rollen (boven murals of einde pagina) ───────────────
            roll_y_max = mural_y_start - 1 if mural_y_start else footer_y

            left_words  = [w for w in words if w['x0'] < 290]
            right_words = [w for w in words if w['x0'] >= 290]

            left_rows  = _parse_rows(left_words,  LEFT_COLS,  roll_y_max)
            right_rows = _parse_rows(right_words, RIGHT_COLS, roll_y_max)

            for raw in left_rows + right_rows:
                item_no = ' '.join(raw.get('item_no', []))
                if not ITEM_RE.match(item_no):
                    continue

                price_m2_f   = _to_float(raw.get('price_m2', []))
                price_roll_f = _to_float(raw.get('price_roll', []))
                width        = ' '.join(t for t in raw.get('width_cm', []) if t != '€')
                length       = ' '.join(t for t in raw.get('length_m', []) if t != '€')

                # Compleet als prijs aanwezig is (samenstelling is n.v.t. voor behang)
                missing = [] if (price_m2_f or price_roll_f) else ['prijs']

                all_items.append({
                    'item_no':         item_no,
                    'article':         current_coll,
                    'collection':      current_coll,
                    'width_cm':        width,
                    'length_m':        length,
                    'repeat_h_cm':     '',
                    'repeat_w_cm':     '',
                    'composition':     '',
                    'martindale':      None,
                    'flame_retardant': False,
                    'coating_cs':      False,
                    'price_excl':      price_roll_f,
                    'price_rrp':       price_m2_f,
                    'colors':          '',
                    'features':        'behang',
                    'missing_fields':  missing,
                })

            # ── Murals ──────────────────────────────────────────────────────
            if mural_y_start:
                ml = [w for w in words if w['x0'] < 290 and w['top'] >= mural_y_start]
                mr = [w for w in words if w['x0'] >= 290 and w['top'] >= mural_y_start]
                for raw in _parse_rows(ml, LEFT_MURAL, footer_y) + \
                           _parse_rows(mr, RIGHT_MURAL, footer_y):
                    item_no = ' '.join(raw.get('item_no', []))
                    if not ITEM_RE.match(item_no):
                        continue
                    price_f = _to_float(raw.get('price_item', []))
                    missing = [] if price_f else ['prijs']
                    all_items.append({
                        'item_no':         item_no,
                        'article':         current_coll,
                        'collection':      current_coll,
                        'width_cm':        '',
                        'length_m':        '',
                        'repeat_h_cm':     '',
                        'repeat_w_cm':     '',
                        'composition':     '',
                        'martindale':      None,
                        'flame_retardant': False,
                        'coating_cs':      False,
                        'price_excl':      price_f,
                        'price_rrp':       None,
                        'colors':          '',
                        'features':        'mural',
                        'missing_fields':  missing,
                    })

    # Dedup op item_no
    seen: set[str] = set()
    unique = []
    for i in all_items:
        if i['item_no'] not in seen:
            seen.add(i['item_no'])
            unique.append(i)

    return _build_result(unique)


# ─── Helpers ─────────────────────────────────────────────────────────────────

def _empty_result() -> dict:
    return {'items': [], 'total_items': 0, 'complete_items': 0,
            'incomplete_items': 0, 'completion_pct': 0, 'gap_items': []}


def _build_result(items: list) -> dict:
    total = len(items)
    complete = sum(1 for i in items if not i.get('missing_fields'))
    pct = round(complete * 100 / total) if total else 0
    gap = [{'item_no': i['item_no'], 'article': i.get('article', ''),
            'missing_fields': i['missing_fields']}
           for i in items if i.get('missing_fields')]
    return {'items': items, 'total_items': total, 'complete_items': complete,
            'incomplete_items': total - complete, 'completion_pct': pct, 'gap_items': gap}
